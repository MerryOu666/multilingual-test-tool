import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def to_excel_bytes(results, lang_name):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{lang_name}测试用例"

    headers = ["序号", "类别", "语种", "语种代码", "测试用例"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, item in enumerate(results, 2):
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=item.get(key, ""))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(key == "测试用例"))

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def to_csv_bytes(results):
    buf = io.StringIO()
    buf.write("﻿")
    headers = ["序号", "类别", "语种", "语种代码", "测试用例"]
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for item in results:
        writer.writerow(item)
    return buf.getvalue().encode("utf-8")


def make_filename(lang_name, ext):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{lang_name}_测试用例_{ts}.{ext}"
